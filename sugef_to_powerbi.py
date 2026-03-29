"""
sugef_to_powerbi.py  v2
=======================
Limpia el reporte de Indicadores Financieros de la SUGEF
y genera archivos listos para Power BI.

Fuente:
    sugef.fi.cr → Servicios → Reportes → Indicadores Financieros
    Seleccionar: Sistema Financiero Nacional → Generar → Excel

Uso:
    python sugef_to_powerbi.py                    # detecta reporte*.xls* automáticamente
    python sugef_to_powerbi.py mi_reporte.xls

Outputs  (en ./output_powerbi/):
    kpi_pivot.csv      ← conectar directo a Power BI
    kpi_largo.csv      ← para análisis en Python/SQL
    kpi_bancario.xlsx  ← Excel formateado con diccionario de indicadores

Requisitos:
    pip install pandas openpyxl
    (xlrd opcional para .xls nativos; si no está, usa LibreOffice)
"""

import logging
import subprocess
import sys
from pathlib import Path

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────────────────────
# CONFIGURACIÓN — editá aquí para agregar/quitar bancos o indicadores
# ─────────────────────────────────────────────────────────────────────────────

BANCOS: dict[str, str] = {
    "BANCO BAC SANJOSE":    "BAC San José",
    "BANCO BCT":            "BCT",
    "BANCO CATHAY":         "Cathay",
    "BANCO CMB":            "CMB",
    "BANCO DAVIVIENDA":     "Davivienda",
    "BANCO DE COSTA RICA":  "BCR",
    "BANCO GENERAL":        "Banco General",
    "BANCO IMPROSA":        "Improsa",
    "BANCO LAFISE":         "Lafise",
    "BANCO NACIONAL":       "Banco Nacional",
    "BANCO POPULAR":        "Banco Popular",
    "BANCO PROMERICA":      "Promerica",
    "SCOTIABANK":           "Scotiabank",
}

# raw SUGEF name → nombre corto para Power BI
INDICADORES: dict[str, str] = {
    "Morosidad mayor a 90 días y cobro judicial / Cartera Directa":
        "mora_90d",
    "Cartera (A+B) / Cartera Total":
        "cartera_ab",
    "Estimaciones sobre cartera de créditos / Cartera con atraso mayor a 90 días y co":
        "cobertura_provisiones",
    "Rentabilidad nominal sobre Patrimonio Promedio":
        "roe",
    "Utilidad Operacional Bruta / Gastos de Administración  1/":
        "eficiencia_op",
    "Activo Productivo / Activo total":
        "activo_productivo_ratio",
    "Captaciones a plazo con el público / Pasivo con costo":
        "captaciones_plazo",
    "Activo Productivo de Intermediación Financiera/ Pasivo con Costo  1/":
        "spread_intermediacion",
}

# Metadatos por KPI: (nombre display, descripción, unidad, interpretación)
KPI_META: dict[str, tuple[str, str, str, str]] = {
    "mora_90d":               ("Mora >90d",           "Cartera con más de 90 días de atraso o en cobro judicial",          "%",     "Menor = mejor"),
    "cartera_ab":             ("Cartera A+B",          "Cartera clasificada en categorías A y B (bajo riesgo SUGEF)",        "%",     "Mayor = mejor"),
    "cobertura_provisiones":  ("Cobertura Prov.",      "Veces que las estimaciones cubren la cartera morosa >90d",           "veces", ">1x = bien cubierto"),
    "roe":                    ("ROE",                  "Rentabilidad nominal sobre patrimonio promedio",                     "%",     "Mayor = más rentable"),
    "eficiencia_op":          ("Eficiencia Op.",       "Veces que la utilidad operativa cubre los gastos administrativos",   "veces", "Mayor = más eficiente"),
    "activo_productivo_ratio":("Activo Prod./Total",   "% del activo total que genera ingresos financieros",                "%",     "Mayor = más eficiente"),
    "captaciones_plazo":      ("Captaciones Plazo",    "% del fondeo proveniente de depósitos a plazo del público",         "%",     "Contextual"),
    "spread_intermediacion":  ("Spread Interm.",       "Activo productivo de intermediación / pasivo con costo",            "veces", "Mayor = mejor"),
}

# KPIs que son porcentajes (valores 0-100) vs los que son "veces"
KPI_PORCENTAJE = {"mora_90d", "cartera_ab", "roe", "activo_productivo_ratio", "captaciones_plazo"}
KPI_VECES      = {"cobertura_provisiones", "eficiencia_op", "spread_intermediacion"}

# Colores Excel
_AZUL_OSCURO = "1F3864"
_AZUL_MEDIO  = "2E75B6"
_VERDE_BAC   = "E2EFDA"
_GRIS_BORDE  = "BDD7EE"

logging.basicConfig(format="%(message)s", level=logging.INFO)
log = logging.getLogger(__name__)


# ─────────────────────────────────────────────────────────────────────────────
# LECTURA
# ─────────────────────────────────────────────────────────────────────────────

def _a_xlsx(ruta: Path) -> Path:
    """Convierte .xls → .xlsx con LibreOffice (fallback cuando xlrd no está)."""
    log.info("  → Convirtiendo .xls con LibreOffice...")
    r = subprocess.run(
        ["libreoffice", "--headless", "--convert-to", "xlsx",
         "--outdir", str(ruta.parent), str(ruta)],
        capture_output=True, text=True, timeout=60,
    )
    if r.returncode != 0:
        raise RuntimeError(f"LibreOffice falló:\n{r.stderr}")
    out = ruta.with_suffix(".xlsx")
    log.info("  ✓ Convertido: %s", out.name)
    return out


def _abrir_workbook(ruta: Path) -> tuple[Path, openpyxl.Workbook]:
    """
    Abre el archivo para lectura con openpyxl.
    Si es .xls, lo convierte a .xlsx primero:
      - con xlrd+pandas si está disponible (funciona en Windows sin LibreOffice)
      - con LibreOffice como fallback (Linux/Mac)
    """
    if ruta.suffix.lower() == ".xls":
        ruta_xlsx = ruta.with_suffix(".xlsx")
        try:
            # xlrd instalado: leer con pandas y re-guardar como xlsx
            df_sheets = pd.read_excel(str(ruta), engine="xlrd", sheet_name=None, header=None)
            with pd.ExcelWriter(str(ruta_xlsx), engine="openpyxl") as writer:
                for nombre, df in df_sheets.items():
                    df.to_excel(writer, sheet_name=nombre, index=False, header=False)
            log.info("  ✓ Convertido con xlrd: %s", ruta_xlsx.name)
        except ImportError:
            # xlrd no disponible: intentar LibreOffice (Linux/Mac)
            ruta_xlsx = _a_xlsx(ruta)
        ruta = ruta_xlsx
    return ruta, openpyxl.load_workbook(str(ruta), read_only=True, data_only=True)


def _detectar_col_fechas(ws) -> tuple[int, dict[int, str]]:
    """
    Recorre las primeras 20 filas buscando la que contiene columnas MM/YYYY.
    Retorna (índice_de_fila_header, {col_index: "MM/YYYY"}).
    """
    for i, fila in enumerate(ws.iter_rows(max_row=20, values_only=True)):
        fechas = {
            j: str(v)
            for j, v in enumerate(fila)
            if v and isinstance(v, str) and len(v) == 7
            and v[2] == "/" and v[:2].isdigit() and v[3:].isdigit()
        }
        if len(fechas) >= 5:
            return i, fechas
    raise ValueError("No se encontró la fila de fechas (MM/YYYY) en el archivo.")


def leer_sugef(ruta: Path) -> pd.DataFrame:
    """
    Lee el XLS de SUGEF y retorna un DataFrame largo con columnas:
        banco, periodo, indicador, valor
    """
    log.info("\n📂 Leyendo: %s", ruta.name)
    ruta, wb = _abrir_workbook(ruta)
    ws = wb.active

    header_idx, col_fechas = _detectar_col_fechas(ws)
    log.info("  Período: %s → %s  (%d meses)",
             list(col_fechas.values())[-1],
             list(col_fechas.values())[0],
             len(col_fechas))

    registros: list[dict] = []
    banco_corto: str | None = None

    for i, fila in enumerate(ws.iter_rows(values_only=True)):
        # Ignorar filas de título/encabezado antes de los datos
        if i <= header_idx:
            continue

        # SUGEF pone el nombre del banco una sola vez; las filas siguientes
        # tienen col[2] vacía. Mantenemos banco_corto como estado y lo
        # actualizamos solo cuando aparece una entidad nueva.
        # Si no está en BANCOS (cooperativas, casas de cambio),
        # banco_corto queda None y sus filas se ignoran automáticamente.
        if fila[2] and fila[3]:
            clave = str(fila[2]).upper()
            banco_corto = next(
                (nombre for patron, nombre in BANCOS.items() if patron in clave),
                None,
            )

        if banco_corto and fila[3]:
            ind_raw = str(fila[3]).strip()
            ind_corto = INDICADORES.get(ind_raw)
            if ind_corto:
                for col_idx, periodo in col_fechas.items():
                    v = fila[col_idx]
                    if v is not None:
                        try:
                            # float() descarta strings vacíos y celdas fusionadas
                            registros.append({
                                "banco":     banco_corto,
                                "periodo":   periodo,
                                "indicador": ind_corto,
                                "valor":     round(float(v), 4),
                            })
                        except (ValueError, TypeError):
                            pass

    wb.close()
    log.info("  ✓ %d registros extraídos", len(registros))
    return pd.DataFrame(registros)


# ─────────────────────────────────────────────────────────────────────────────
# TRANSFORMACIÓN
# ─────────────────────────────────────────────────────────────────────────────

def transformar(df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    Convierte el DataFrame largo en:
    - df_largo : limpio, con columna `fecha` datetime
    - df_pivot : una columna por KPI, listo para Power BI
    """
    log.info("\n🧹 Transformando...")

    # "03/2025" → datetime(2025,3,1): Power BI necesita fecha real para el eje de tiempo.
    # errors="coerce" convierte períodos malformados en NaT en vez de romper.
    df["fecha"] = pd.to_datetime(
        "01/" + df["periodo"], format="%d/%m/%Y", errors="coerce"
    )

    df = (
        df
        .dropna(subset=["fecha", "valor"])            # elimina NaT y valores nulos
        .drop_duplicates(subset=["banco", "fecha", "indicador"])  # mismo indicador con nombre levemente distinto entre períodos
        .sort_values(["banco", "indicador", "fecha"]) # orden cronológico para que PBI grafique sin configuración extra
        .reset_index(drop=True)                       # renumera índice tras eliminar filas
    )

    df_largo = df[["banco", "periodo", "fecha", "indicador", "valor"]].copy()

    # Formato long → wide: una columna por KPI, una fila por banco+período.
    # rename_axis limpia el nombre "indicador" que pandas deja en el eje de columnas.
    df_pivot = (
        df
        .pivot(index=["banco", "periodo", "fecha"], columns="indicador", values="valor")
        .reset_index()
        .rename_axis(None, axis=1)
    )

    # Si un banco no reportó algún KPI en ningún período, la columna no existe
    # tras el pivot. La creamos con NA para que el CSV tenga siempre 11 columnas.
    for kpi in INDICADORES.values():
        if kpi not in df_pivot.columns:
            df_pivot[kpi] = pd.NA

    cols_kpi = list(INDICADORES.values())
    df_pivot = df_pivot[["banco", "periodo", "fecha"] + cols_kpi]

    log.info("  ✓ largo  → %d filas", len(df_largo))
    log.info("  ✓ pivot  → %d filas × %d cols", *df_pivot.shape)
    log.info("  ✓ bancos → %s", sorted(df_pivot["banco"].unique()))
    return df_largo, df_pivot


# ─────────────────────────────────────────────────────────────────────────────
# EXPORT — CSV
# ─────────────────────────────────────────────────────────────────────────────

def exportar_csv(df_largo: pd.DataFrame, df_pivot: pd.DataFrame, carpeta: Path) -> None:
    for df, nombre in [(df_largo, "kpi_largo.csv"), (df_pivot, "kpi_pivot.csv")]:
        ruta = carpeta / nombre
        df.to_csv(ruta, index=False, encoding="utf-8-sig")
        log.info("   %s  (%d KB)", ruta.name, ruta.stat().st_size // 1024)


# ─────────────────────────────────────────────────────────────────────────────
# EXPORT — EXCEL
# ─────────────────────────────────────────────────────────────────────────────

def _estilo_celda(ws, fila: int, col: int, **kwargs) -> openpyxl.cell.Cell:
    """Aplica estilos a una celda y la retorna."""
    c = ws.cell(row=fila, column=col)
    _borde = Border(*[Side(style="thin", color=_GRIS_BORDE)] * 4)
    c.border = _borde
    for attr, val in kwargs.items():
        setattr(c, attr, val)
    return c


def _escribir_hoja_kpis(ws, df_pivot: pd.DataFrame) -> None:
    """Escribe la hoja de datos KPIs con formato."""
    cols = list(df_pivot.columns)
    n_cols = len(cols)

    DISPLAY: dict[str, str] = {k: v[0] for k, v in KPI_META.items()} | {
        "banco": "Banco", "periodo": "Período", "fecha": "Fecha"
    }
    ANCHOS: dict[str, int] = {
        "banco": 18, "periodo": 9, "fecha": 12,
        **{k: 14 for k in INDICADORES.values()},
    }
    NUMFMT: dict[str, str] = {
        **{k: "0.00" for k in KPI_PORCENTAJE},
        **{k: '0.00"x"' for k in KPI_VECES},
    }

    fill_header  = PatternFill("solid", fgColor=_AZUL_OSCURO)
    fill_subh    = PatternFill("solid", fgColor=_AZUL_MEDIO)
    fill_bac     = PatternFill("solid", fgColor=_VERDE_BAC)
    font_white   = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    font_data    = Font(name="Calibri", size=9)
    font_bold    = Font(name="Calibri", size=9, bold=True)
    align_center = Alignment(horizontal="center", vertical="center")
    align_right  = Alignment(horizontal="right",  vertical="center")
    align_left   = Alignment(horizontal="left",   vertical="center")

    # Fila 1: título fusionado
    ws.merge_cells(f"A1:{get_column_letter(n_cols)}1")
    t = ws["A1"]
    t.value = "KPIs Sistema Financiero Costarricense — Fuente: SUGEF"
    t.font = Font(name="Calibri", bold=True, size=13, color="FFFFFF")
    t.fill = fill_header
    t.alignment = align_center
    ws.row_dimensions[1].height = 28

    # Fila 2: encabezados de columna
    for j, col in enumerate(cols, 1):
        c = ws.cell(row=2, column=j,
                    value=DISPLAY.get(col, col))
        c.font, c.fill, c.alignment = font_white, fill_subh, align_center
        c.border = Border(*[Side(style="thin", color=_GRIS_BORDE)] * 4)
    ws.row_dimensions[2].height = 30

    # Datos: ws.append() es ~20× más rápido que cell-by-cell
    es_bac: list[bool] = []
    for row in df_pivot.itertuples(index=False):
        fecha_val = row.fecha
        fila_vals = []
        for col in cols:
            v = getattr(row, col)
            if col == "fecha":
                fila_vals.append(fecha_val.strftime("%Y-%m-%d") if pd.notna(fecha_val) else None)
            else:
                fila_vals.append(None if pd.isna(v) else v)
        ws.append(fila_vals)
        es_bac.append(str(row.banco) == "BAC San José")

    # Estilos post-escritura (un único recorrido sobre filas de datos)
    borde = Border(*[Side(style="thin", color=_GRIS_BORDE)] * 4)
    for rel_idx, bac in enumerate(es_bac):
        fila_excel = rel_idx + 3
        for j, col in enumerate(cols, 1):
            c = ws.cell(row=fila_excel, column=j)
            c.border = borde
            c.font   = font_bold if col == "banco" else font_data
            c.alignment = align_left if j <= 2 else align_right
            if bac:
                c.fill = fill_bac
            if fmt := NUMFMT.get(col):
                c.number_format = fmt

    # Anchos y freeze
    for j, col in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(j)].width = ANCHOS.get(col, 14)
    ws.freeze_panes = "D3"


def _escribir_hoja_diccionario(ws, wb: openpyxl.Workbook) -> None:
    """Escribe la hoja de diccionario de indicadores."""
    fill_h = PatternFill("solid", fgColor=_AZUL_OSCURO)
    borde  = Border(*[Side(style="thin", color=_GRIS_BORDE)] * 4)

    headers = ["Columna CSV", "Nombre Display", "Descripción", "Unidad", "Interpretación"]
    for j, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=j, value=h)
        c.font = Font(bold=True, color="FFFFFF", name="Calibri")
        c.fill = fill_h
        c.border = borde

    for i, (kpi, (display, desc, unidad, interp)) in enumerate(KPI_META.items(), 2):
        for j, val in enumerate([kpi, display, desc, unidad, interp], 1):
            c = ws.cell(row=i, column=j, value=val)
            c.font   = Font(name="Calibri", size=9)
            c.border = borde

    for col, ancho in zip("ABCDE", [22, 22, 60, 10, 28]):
        ws.column_dimensions[col].width = ancho


def exportar_excel(df_pivot: pd.DataFrame, carpeta: Path) -> None:
    """Genera kpi_bancario.xlsx con hoja de datos y diccionario."""
    wb = openpyxl.Workbook()

    ws_kpi = wb.active
    ws_kpi.title = "KPIs"
    _escribir_hoja_kpis(ws_kpi, df_pivot)

    ws_dic = wb.create_sheet("Diccionario")
    _escribir_hoja_diccionario(ws_dic, wb)

    ruta = carpeta / "kpi_bancario.xlsx"
    wb.save(str(ruta))
    log.info("   %s  (%d KB)", ruta.name, ruta.stat().st_size // 1024)


# ─────────────────────────────────────────────────────────────────────────────
# RESUMEN CONSOLA
# ─────────────────────────────────────────────────────────────────────────────

def _resumen(df_pivot: pd.DataFrame) -> None:
    periodo = df_pivot["periodo"].max()
    df = df_pivot[df_pivot["periodo"] == periodo]
    sep = "═" * 58

    log.info("\n%s", sep)
    log.info("  KPIs al %s", periodo)
    log.info("%s", sep)

    for kpi, label, asc in [("mora_90d", "Mora >90d (%)", True),
                             ("roe",      "ROE (%)",       False)]:
        log.info("\n  %s", label)
        for _, r in df.sort_values(kpi, ascending=asc).iterrows():
            marca = " ← BAC" if r["banco"] == "BAC San José" else ""
            log.info("    %-20s %6.2f%s", r["banco"], r[kpi] or 0, marca)

    log.info("\n%s\n", sep)


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────

def main() -> None:
    if len(sys.argv) > 1:
        ruta = Path(sys.argv[1])
    else:
        candidatos = sorted(Path(".").glob("reporte*.xls*"))
        if not candidatos:
            log.error("No se encontró ningún archivo reporte*.xls*")
            log.error("Uso: python sugef_to_powerbi.py archivo.xls")
            sys.exit(1)
        ruta = candidatos[0]
        log.info("🔍 Archivo detectado: %s", ruta.name)

    if not ruta.exists():
        log.error("Archivo no encontrado: %s", ruta)
        sys.exit(1)

    carpeta = ruta.parent / "output_powerbi"
    carpeta.mkdir(exist_ok=True)

    df_raw             = leer_sugef(ruta)
    df_largo, df_pivot = transformar(df_raw)

    log.info("\n💾 Exportando...")
    exportar_csv(df_largo, df_pivot, carpeta)
    exportar_excel(df_pivot, carpeta)

    _resumen(df_pivot)

    log.info("✅ Listo → %s", carpeta.resolve())
    log.info("\n   Power BI: Obtener datos → CSV → kpi_pivot.csv")
    log.info("   La columna 'fecha' viene en YYYY-MM-DD ✓\n")


if __name__ == "__main__":
    main()
